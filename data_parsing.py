import re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# 1. Text extraction

FOOTER_PATTERNS = [
    re.compile(r'Current version for .+? to date \(accessed .+?\)'),
    re.compile(r'Certified by the NSW Parliamentary Counsel.+'),
    re.compile(r'Residential Tenancies Act 2010 No 42 \[NSW\]'),
    re.compile(r'Residential T enancies Act 2010 No 42 \[NSW\]'),
    re.compile(r'Page \d+ of \d+'),
]

# Lines from the table of contents / index end with dots then a page number
INDEX_LINE_RE = re.compile(r'\.{3,}\s*\d+\s*$')

def is_index_line(line):
    return bool(INDEX_LINE_RE.search(line))

def clean_line(line):
    for pat in FOOTER_PATTERNS:
        if pat.search(line):
            return None
    return line


def extract_full_text(pdf_path):
    """Extract text from PDF, strip footers, return (full_text, page_positions).
    page_positions: list of (char_offset, footer_page_number)."""
    full_text = ""
    page_positions = []
    page_re = re.compile(r'Page (\d+) of \d+')

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            raw = page.extract_text() or ""
            m = page_re.search(raw)
            footer_page = int(m.group(1)) if m else None

            kept = []
            for line in raw.split('\n'):
                if is_index_line(line):
                    continue
                cleaned = clean_line(line)
                if cleaned is not None:
                    kept.append(cleaned)

            page_text = '\n'.join(kept)
            page_positions.append((len(full_text), footer_page))
            full_text += page_text + '\n'

    return full_text, page_positions


def page_for_offset(offset, page_positions):
    result = page_positions[0][1]
    for start, pnum in page_positions:
        if pnum and offset >= start:
            result = pnum
    return result


# 2. Pre-processing: join wrapped continuation lines for headings
#
# In the raw PDF text, long Part/Schedule titles wrap like:
#   "    Part 3 Provision consequent on enactment of Statute Law"
#   "      (Miscellaneous Provisions) Act (No 2) 2012"
# or Schedule section reference wraps:
#   "    Schedule 1 Membership and procedure of Rental Bond Board"
#   "                                                (Section 178(3))"
# We join these so the heading is on one line.

def join_wrapped_headings(text):
    """Join continuation lines that belong to a Part/Schedule/Division heading."""
    lines = text.split('\n')
    out = []
    i = 0
    heading_re = re.compile(r'^\s{0,8}(Part|Division|Schedule|Subdivision)\s+\d+')
    # A continuation line: indented more than the heading, starts with ( or lowercase
    # or is heavily right-padded (the "(Section 178(3))" case)
    continuation_re = re.compile(r'^\s{6,}(\(|[a-z])')

    while i < len(lines):
        line = lines[i]
        if heading_re.match(line):
            # Look ahead for continuation
            combined = line.rstrip()
            while i + 1 < len(lines):
                nxt = lines[i + 1]
                stripped = nxt.strip()
                if not stripped:
                    break
                # Continuation: deeply indented AND not a new structural keyword
                if (continuation_re.match(nxt)
                        and not heading_re.match(nxt)
                        and not re.match(r'^\s{0,8}\d+[A-Z]{0,2}\s+[A-Z(]', nxt)):
                    combined = combined + ' ' + stripped
                    i += 1
                else:
                    break
            out.append(combined)
        else:
            out.append(line)
        i += 1

    return '\n'.join(out)


# 3. Structural patterns

# "Part 1 Preliminary" / "Part 13" (bare) / "Part 13 (Repealed)"
# Bare Part (no title after number) handled separately
PART_RE      = re.compile(r'(?m)^\s{0,8}(Part\s+(\d+[A-Z]?)(?:\s+\S[^\n]*)?)$')
BARE_PART_RE = re.compile(r'(?m)^\s{0,8}Part\s+(\d+[A-Z]?)\s*$')

DIV_RE       = re.compile(r'(?m)^\s{0,8}(Division\s+\d+[A-Z]?\s+\S[^\n]+)$')
SUBDIV_RE    = re.compile(r'(?m)^\s{0,8}(Subdivision\s+\d+[A-Z]?\s+\S[^\n]+)$')

# "Schedule 1 ..."  /  "Schedule 3 (Repealed)"
SCHED_RE     = re.compile(r'(?m)^\s{0,8}(Schedule\s+(\d+[A-Z]?)\s*\S[^\n]*)$')
BARE_SCHED_RE= re.compile(r'(?m)^\s{0,8}Schedule\s+(\d+[A-Z]?)\s*\(Repealed\)\s*$')

# Range-repealed line: "228A–228C, 230 (Repealed)"  (en-dash or hyphen)
RANGE_REPEAL_RE = re.compile(r'(?m)^\s{0,8}(\d+[A-Z]{0,2}[\u2013\-]\d+[A-Z]{0,2}(?:,\s*\d+[A-Z]{0,2})*)\s+\(Repealed\)\s*$')

# Normal section: exactly 4-space indent distinguishes real headings from
# definition lines like '1977 Act means...' which are indented 8+ spaces.
# Allow 3-5 spaces to absorb minor PDF extraction jitter.
SECTION_RE   = re.compile(r'(?m)^\s{0,8}(\d{1,3}[A-Z]{0,5})\s+([A-Z(][^\n]{1,120})$')


def is_structural(line):
    """True if line is a Part / Division / Subdivision / Schedule heading."""
    s = line.strip()
    return bool(
        re.match(r'Part\s+\d+', s)
        or re.match(r'Division\s+\d+', s)
        or re.match(r'Subdivision\s+\d+', s)
        or re.match(r'Schedule\s+\d+', s)
    )


# 4. Build timelines

def join_continuation(full_text, match_end):
    """If the line immediately after a heading match is a continuation
    (indented, does not look like a section/structural line, starts with
    a parenthesis or lowercase/mid-word), absorb it into the label."""
    rest = full_text[match_end:]
    line_m = re.match(r'\n[ \t]*([^\n]+)', rest)
    if not line_m:
        return ""
    candidate = line_m.group(1).strip()
    if re.match(r'[a-z(]', candidate) and not is_structural(candidate):
        return " " + candidate
    return ""


def build_timelines(full_text):
    """Return sorted timelines for Part, Division, Subdivision, Schedule."""
    part_tl = []
    for m in PART_RE.finditer(full_text):
        label = m.group(1).strip()
        # Absorb a wrapped continuation line (e.g. "(Miscellaneous Provisions) Act (No 2) 2012")
        label += join_continuation(full_text, m.end())
        part_tl.append((m.start(), label))

    div_tl = [(m.start(), m.group(1).strip()) for m in DIV_RE.finditer(full_text)]
    subdiv_tl = [(m.start(), m.group(1).strip()) for m in SUBDIV_RE.finditer(full_text)]
    sched_tl = [(m.start(), m.group(1).strip(), m.group(2)) for m in SCHED_RE.finditer(full_text)]

    return part_tl, div_tl, subdiv_tl, sched_tl


def latest_before(timeline, offset):
    """Return the most recent value from a (pos, value, ...) timeline before offset."""
    result = None
    for item in timeline:
        pos = item[0]
        val = item[1]
        if pos < offset:
            result = val
    return result


def context_at(offset, part_tl, div_tl, subdiv_tl, sched_tl):
    """Return (part, division, subdivision, schedule_label, schedule_number)."""
    part   = latest_before(part_tl,   offset)
    div    = latest_before(div_tl,    offset)
    subdiv = latest_before(subdiv_tl, offset)

    sched_label = None
    sched_num   = None
    for pos, label, num in sched_tl:
        if pos < offset:
            sched_label = label
            sched_num   = num

    # Determine the most recent top-level structural change
    last_part_pos   = max((p for p, _ in part_tl   if p < offset), default=-1)
    last_div_pos    = max((p for p, _ in div_tl    if p < offset), default=-1)
    last_subdiv_pos = max((p for p, _ in subdiv_tl if p < offset), default=-1)
    last_sched_pos  = max((p for p, _, __ in sched_tl if p < offset), default=-1)

    # When inside a Schedule: clear main-Act division/subdivision that predate the schedule
    if last_sched_pos > last_part_pos:
        if last_div_pos < last_sched_pos:
            div = None
        if last_subdiv_pos < last_sched_pos:
            subdiv = None

    # When Part changes, clear division and subdivision if they predate the new part
    if last_part_pos > last_div_pos:
        div = None
    if last_part_pos > last_subdiv_pos:
        subdiv = None

    # When Division changes, clear subdivision if it predates the new division
    if last_div_pos > last_subdiv_pos:
        subdiv = None

    return part, div, subdiv, sched_label, sched_num


# 5. Section index (normal sections, namespace by schedule)

def build_section_index(full_text, sched_tl):
    """Return sorted list of (start_offset, sid, title).
    Key: (schedule_number_or_None, sid) to prevent collision when
    section numbers reset inside each Schedule."""
    seen = set()
    sections = []

    for m in SECTION_RE.finditer(full_text):
        sid   = m.group(1)
        title = m.group(2).strip()
        start = m.start()

        if is_structural(m.group(0)):
            continue

        cur_sched_num = None
        for pos, _, num in sched_tl:
            if pos < start:
                cur_sched_num = num

        key = (cur_sched_num, sid)
        if key in seen:
            continue
        seen.add(key)

        sections.append((start, sid, title))

    sections.sort(key=lambda x: x[0])
    return sections


# 6. Synthetic records for structural elements with no matchable sections
#    a) Bare "Part N" with a range-repeal line ("228A-228C, 230 (Repealed)")
#    b) Parts that contain no sections at all
#    c) Schedules that are "(Repealed)" with no content

def build_synthetic_records(full_text, section_index, part_tl, sched_tl,
                             page_positions, div_tl, subdiv_tl,
                             pdf_filename="", last_updated=""):
    section_offsets = {s for s, _, _ in section_index}
    records = []

    # --- Bare Part N (no title on same line) + range-repeal annotation ---
    for m in BARE_PART_RE.finditer(full_text):
        part_start = m.start()
        part_num   = m.group(1)
        part_label = f"Part {part_num}"

        # Grab the next non-blank line — it should be the range-repeal annotation
        rest = full_text[m.end():]
        annotation = ""
        for line in rest.split('\n'):
            stripped = line.strip()
            if stripped:
                annotation = stripped
                break

        # Check no real section falls right after (within ~200 chars)
        # If annotation matches range-repeal pattern, use it as title/content
        range_m = RANGE_REPEAL_RE.search('\n' + annotation)
        if range_m or annotation.lower().startswith('(repealed)'):
            title   = annotation if annotation else "(Repealed)"
            content = ""
        else:
            # Part has normal content — skip, will be picked up by inject below
            continue

        page = page_for_offset(part_start, page_positions)
        part, div, subdiv, sched_label, sched_num = context_at(
            part_start, part_tl, div_tl, subdiv_tl, sched_tl
        )
        records.append({
            "section":         part_label,
            "title":           title,
            "section_content": content,
            "page":            page,
            "schedule":        sched_label,
            "type":            "legislation",
            "pdf_filename":    pdf_filename,
            "jurisdiction":    "NSW",
            "part":            part_label,
            "division":        None,
            "subdivision":     None,
            "last_updated":    last_updated,
            "_sort_key":       part_start,
        })

    # --- Parts with no sections (not bare) ---
    for i, (part_start, part_label) in enumerate(part_tl):
        # Skip bare parts — already handled above
        bare = BARE_PART_RE.match(
            full_text[part_start:full_text.index('\n', part_start) + 1]
        )
        if bare:
            continue

        next_part_start = part_tl[i + 1][0] if i + 1 < len(part_tl) else len(full_text)
        has_section = any(part_start < s < next_part_start for s in section_offsets)

        if not has_section and '(Repealed)' in part_label:
            page = page_for_offset(part_start, page_positions)
            part, div, subdiv, sched_label, sched_num = context_at(
                part_start, part_tl, div_tl, subdiv_tl, sched_tl
            )
            records.append({
                "section":         part_label,
                "title":           "(Repealed)",
                "section_content": "",
                "page":            page,
                "schedule":        sched_label,
                "type":            "legislation",
                "pdf_filename":    pdf_filename,
                "jurisdiction":    "NSW",
                "part":            part_label,
                "division":        None,
                "subdivision":     None,
                "last_updated":    last_updated,
                "_sort_key":       part_start,
            })

    # --- Repealed Schedules (e.g. "Schedule 3 (Repealed)") ---
    for pos, label, num in sched_tl:
        if '(Repealed)' in label:
            sched_end = len(full_text)
            for p2, l2, n2 in sched_tl:
                if p2 > pos:
                    sched_end = p2
                    break
            has_section = any(pos < s < sched_end for s in section_offsets)
            if not has_section:
                page = page_for_offset(pos, page_positions)
                part, div, subdiv, _, _ = context_at(
                    pos, part_tl, div_tl, subdiv_tl, sched_tl
                )
                records.append({
                    "section":         label,
                    "title":           "(Repealed)",
                    "section_content": "",
                    "page":            page,
                    "schedule":        label,
                    "type":            "legislation",
                    "pdf_filename":    pdf_filename,
                    "jurisdiction":    "NSW",
                    "part":            None,
                    "division":        None,
                    "subdivision":     None,
                    "last_updated":    last_updated,
                    "_sort_key":       pos,
                })

    return records


# 7. Content extraction helper

def extract_content(full_text, heading_end, end):
    raw = full_text[heading_end:end].strip()
    cleaned = [l for l in raw.split('\n') if not is_structural(l)]
    return re.sub(r'\n{3,}', '\n\n', '\n'.join(cleaned)).strip()


# 8. Main parse

def parse_sections(pdf_path, pdf_filename="", last_updated=""):
    raw_text, page_positions = extract_full_text(pdf_path)
    full_text = join_wrapped_headings(raw_text)

    part_tl, div_tl, subdiv_tl, sched_tl = build_timelines(full_text)
    section_index = build_section_index(full_text, sched_tl)

    synthetic = build_synthetic_records(
        full_text, section_index, part_tl, sched_tl,
        page_positions, div_tl, subdiv_tl,
        pdf_filename=pdf_filename, last_updated=last_updated
    )

    records = []

    # Normal section records
    for i, (start, sid, title) in enumerate(section_index):
        end         = section_index[i + 1][0] if i + 1 < len(section_index) else len(full_text)
        heading_end = full_text.index('\n', start) + 1
        content     = extract_content(full_text, heading_end, end)
        page        = page_for_offset(start, page_positions)

        part, div, subdiv, sched_label, sched_num = context_at(
            start, part_tl, div_tl, subdiv_tl, sched_tl
        )

        try:
            section_num = int(sid)
        except ValueError:
            section_num = sid  # e.g. "8A", "87E", "228C"

        records.append({
            "section":         section_num,
            "title":           title,
            "section_content": content,
            "page":            page,
            "schedule":        sched_label,
            "type":            "legislation",
            "pdf_filename":    pdf_filename,
            "jurisdiction":    "NSW",
            "part":            part,
            "division":        div,
            "subdivision":     subdiv,
            "last_updated":    last_updated,
            "_sort_key":       start,
        })

    records.extend(synthetic)
    records.sort(key=lambda r: r["_sort_key"])
    for r in records:
        del r["_sort_key"]

    return records


# 10. Batch: process all PDFs in a folder

def parse_folder(file_configs, output_path):
    """Parse a list of PDFs and write all records to one Excel file.

    file_configs: list of dicts, each with keys:
        pdf_path     - path to the PDF file
        last_updated - version date string for that file (e.g. "15 August 2025")
    output_path: path for the output Excel file
    """
    from pathlib import Path

    all_records = []
    for cfg in file_configs:
        pdf_path     = cfg["pdf_path"]
        last_updated = cfg.get("last_updated", "")
        pdf_filename = Path(pdf_path).name

        print(f"Processing {pdf_filename} ...")
        try:
            records = parse_sections(
                pdf_path,
                pdf_filename=pdf_filename,
                last_updated=last_updated,
            )
            all_records.extend(records)
            print(f"  -> {len(records)} sections extracted")
        except Exception as e:
            print(f"  ERROR processing {pdf_filename}: {e}")

    if all_records:
        save_to_excel(all_records, output_path)
        print(f"\nTotal: {len(all_records)} sections from {len(file_configs)} PDFs -> {output_path}")
    else:
        print("No records extracted.")




def save_to_excel(records, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sections"

    headers = ["Section", "Title", "Section Content", "Metadata",]

    header_fill  = PatternFill("solid", start_color="1F3864")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    alt_fill     = PatternFill("solid", start_color="EEF2FF")
    data_font    = Font(name="Arial", size=9)
    WRAP_COLS    = {3, 13}   # Section Content + Metadata columns
    wrap_align   = Alignment(vertical="top", wrap_text=True)
    nowrap_align = Alignment(vertical="top", wrap_text=False)

    for row_idx, r in enumerate(records, 2):
        # --- Build metadata dict ---
        # Normalise last_updated to YYYY-MM if possible, else keep as-is
        raw_date = r.get("last_updated", "")
        try:
            from datetime import datetime
            for fmt in ("%d %B %Y", "%d %b %Y", "%B %Y", "%b %Y"):
                try:
                    parsed = datetime.strptime(raw_date.strip(), fmt)
                    raw_date = parsed.strftime("%Y-%m")
                    break
                except ValueError:
                    continue
        except Exception:
            pass

        # Strip path/extension from pdf_filename for a clean display name
        from pathlib import Path
        pdf_display = Path(r.get("pdf_filename", "")).stem  # e.g. "Residential Tenancies Act 2010"

        metadata = {
            "type":         r.get("type") or None,
            "pdf_file":     pdf_display or None,
            "part":         r.get("part") or None,
            "division":     r.get("division") or None,
            "subdivision":  r.get("subdivision") or None,
            "schedule":     r.get("schedule") or None,
            "page":         r.get("page") or None,
            "jurisdiction": r.get("jurisdiction") or None,
            "last_updated": raw_date or None,
        }
        metadata_str = json.dumps(metadata, ensure_ascii=False)

        row_data = [
            r["section"], r["title"], r["section_content"],
            metadata_str,   
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=value)
            cell.font      = data_font
            cell.border    = border
            cell.alignment = wrap_align if col in WRAP_COLS else nowrap_align
            if fill:
                cell.fill = fill

    col_widths = [12, 40, 80, 8, 45, 14, 45, 14, 40, 35, 35, 16, 60]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Saved {len(records)} records -> {output_path}")

# 11. Entry point

if __name__ == "__main__":

    # Configure your PDFs here — add one entry per file
    file_configs = [
        {
            "pdf_path":     "no index/anti_discrimination_act_1977.pdf",      # <- path to PDF
            "last_updated": "1 July 2024",     # <- version date for this file
        },
        {
            "pdf_path":     "no index/boarding_house_act_2012_&_regulation_2013.pdf",
            "last_updated": "13 January 2023", 
        },
        {
            "pdf_path":     "no index/c&a_tribunal_act_2013.pdf",      # <- path to PDF
            "last_updated": "28 March 2026",     # <- version date for this file
        },
        {
            "pdf_path":     "no index/property_stock_agents_act_2002_&_reg_2022.pdf",      # <- path to PDF
            "last_updated": "13 MAy 2024",     # <- version date for this file
        },
        {
            "pdf_path":     "no index/residential_2010.pdf",     
            "last_updated": "15 August 2025",   
        },
         {
            "pdf_path":     "no index/residential_regulation_2019.pdf",     
            "last_updated": "13 February 2026",   
        },
         {
            "pdf_path":     "no index/residential(land_lease)_act_2013.pdf",     
            "last_updated": "11 December 2024",   
        },
         {
            "pdf_path":     "no index/strata_act_2015.pdf",     
            "last_updated": "1 April 2026",   
        }

    ]

    output_path = "acts_regulations.xlsx"  # <- output Excel file

    parse_folder(file_configs, output_path)